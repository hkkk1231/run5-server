import openpyxl
import os
import re
import json
from datetime import datetime, date
from openpyxl.styles import Alignment, NamedStyle

# 使用统一的绝对路径配置
from paths import CURRENT_MILEAGE_FILE, EXCEL_SOURCE_FILE, EXCEL_FILE

def extract_numeric_value(text):
    """
    从文本中提取数字，支持多种货币符号和格式
    """
    if text is None:
        return 0.0
    
    # 转换为字符串并移除常见货币符号和空格
    text = str(text).replace('￥', '').replace('¥', '').replace('$', '').replace(',', '').strip()
    
    # 使用正则表达式提取数字（包括小数点）
    match = re.search(r'[\d.]+', text)
    if match:
        try:
            return float(match.group())
        except ValueError:
            return 0.0
    return 0.0

def process_excel_file(input_path, output_path):
    """
    处理Excel文件，正确处理合并单元格和价格计算
    """
    # 加载工作簿，使用data_only=True获取公式的计算结果而不是公式本身
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    worksheet = workbook.active
    
    # 获取表头
    headers = [cell.value for cell in worksheet[1]]
    
    # 找到价格列（表头为"价格"）
    price_col_idx = None
    for idx, header in enumerate(headers):
        if header == "价格":
            price_col_idx = idx
            break
    
    if price_col_idx is None:
        raise ValueError("未找到价格列，请检查表头是否包含'价格'")
    
    # 获取所有合并单元格信息
    merged_ranges = []
    if hasattr(worksheet, 'merged_cells') and worksheet.merged_cells:
        merged_ranges = list(worksheet.merged_cells.ranges)
    
    # 创建新的工作簿用于存储结果
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.title = "Sheet1"
    
    # 创建货币样式
    currency_style = NamedStyle(name="currency")
    currency_style.number_format = '￥#,##0.00'
    if "currency" not in new_workbook.named_styles:
        new_workbook.add_named_style(currency_style)
    
    # 创建文本样式
    text_style = NamedStyle(name="text")
    text_style.number_format = '@'
    if "text" not in new_workbook.named_styles:
        new_workbook.add_named_style(text_style)
    
    # 复制表头并设置格式
    for col_idx, header in enumerate(headers, 1):
        cell = new_worksheet.cell(row=1, column=col_idx, value=header)
        cell.style = text_style
        cell.alignment = Alignment(horizontal='center')
    
    # 添加"当前里程"列到表头
    current_mileage_col_idx = len(headers) + 1
    current_mileage_cell = new_worksheet.cell(row=1, column=current_mileage_col_idx, value="当前里程")
    current_mileage_cell.style = text_style
    current_mileage_cell.alignment = Alignment(horizontal='center')
    
    # 获取原始数据范围
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    # 首先复制所有原始数据到新工作表
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            original_value = worksheet.cell(row=row_idx, column=col_idx).value
            cell = new_worksheet.cell(row=row_idx, column=col_idx, value=original_value)
            cell.style = text_style
            cell.alignment = Alignment(horizontal='center')
        
        # 为新添加的"当前里程"列设置默认值
        current_mileage_cell = new_worksheet.cell(row=row_idx, column=len(headers) + 1, value="None")
        current_mileage_cell.style = text_style
        current_mileage_cell.alignment = Alignment(horizontal='center')
    
    # 处理合并单元格
    for merged_range in merged_ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        merged_max_row, merged_max_col = merged_range.max_row, merged_range.max_col
        
        # 获取合并区域的值（左上角单元格的值）
        cell_value = worksheet.cell(row=min_row, column=min_col).value
        
        # 如果是价格列且是合并单元格，需要平分价格
        if price_col_idx is not None and min_col == price_col_idx + 1:
            # 计算合并的行数
            rows_count = merged_max_row - min_row + 1
            
            # 提取价格数值
            price_value = extract_numeric_value(cell_value)
            
            # 平分价格
            if rows_count > 0 and price_value > 0:
                divided_price = price_value / rows_count
                
                # 为合并区域中的每个单元格设置平分后的价格
                for r in range(min_row, merged_max_row + 1):
                    cell = new_worksheet.cell(row=r, column=min_col, value=divided_price)
                    cell.style = currency_style
                    cell.alignment = Alignment(horizontal='center')
        else:
            # 非价格列的合并单元格，为每个单元格填充相同的值
            for r in range(min_row, merged_max_row + 1):
                for c in range(min_col, merged_max_col + 1):
                    cell = new_worksheet.cell(row=r, column=c, value=cell_value)
                    cell.style = text_style
                    cell.alignment = Alignment(horizontal='center')
    
    # 确保价格列所有单元格都有正确的格式
    for row_idx in range(2, max_row + 1):
        cell = new_worksheet.cell(row=row_idx, column=price_col_idx + 1)
        if cell.value is not None:
            # 如果是数字但不是货币格式，设置为货币格式
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, str):
                cell.style = currency_style
            # 如果是字符串但不是以￥开头，转换为数字再设置货币格式
            elif isinstance(cell.value, str) and not cell.value.startswith('￥'):
                numeric_value = extract_numeric_value(cell.value)
                if numeric_value > 0:
                    cell.value = numeric_value
                    cell.style = currency_style
        cell.alignment = Alignment(horizontal='center')
    
    # 加载当前里程数据
    with open(str(CURRENT_MILEAGE_FILE), 'r', encoding='utf-8') as f:
        mileage_data = json.load(f)
    
    # 获取各列的索引
    headers = [cell.value for cell in new_worksheet[1]]
    student_id_col_idx = headers.index("学号") + 1  # +1 因为openpyxl列索引从1开始
    password_col_idx = headers.index("密码") + 1
    path_col_idx = headers.index("途径") + 1
    less_than_4_col_idx = headers.index("<4") + 1
    target_mileage_col_idx = headers.index("目标里程") + 1
    manual_stop_col_idx = headers.index("手动停止") + 1
    # "当前里程"列是我们刚刚添加的，所以它的索引是headers的长度
    current_mileage_col_idx = len(headers)
    
    # 处理每一行数据
    for row_idx in range(2, max_row + 1):
        # 获取学号并确保为字符串类型，与JSON文件中的键类型一致
        student_id = str(new_worksheet.cell(row=row_idx, column=student_id_col_idx).value)
        
        # 密码处理：如果密码列为空，则使用学号作为密码
        password_cell = new_worksheet.cell(row=row_idx, column=password_col_idx)
        if password_cell.value is None or password_cell.value == "":
            password_cell.value = student_id
        
        # <4字段处理：如果为空则填充默认值，如果途径列值为"追逐"，则<4列设为1，否则为0
        path_value = new_worksheet.cell(row=row_idx, column=path_col_idx).value
        less_than_4_cell = new_worksheet.cell(row=row_idx, column=less_than_4_col_idx)
        if less_than_4_cell.value is None or less_than_4_cell.value == "":
            less_than_4_cell.value = 1 if path_value == "追逐" else 0
        
        # 设置目标里程和手动停止列的固定值
        target_mileage_cell = new_worksheet.cell(row=row_idx, column=target_mileage_col_idx)
        if target_mileage_cell.value is None or target_mileage_cell.value == "":
            target_mileage_cell.value = 60
            
        manual_stop_cell = new_worksheet.cell(row=row_idx, column=manual_stop_col_idx)
        if manual_stop_cell.value is None or manual_stop_cell.value == "":
            manual_stop_cell.value = 0
        
        # 添加当前里程数据
        if student_id in mileage_data:
            new_worksheet.cell(row=row_idx, column=current_mileage_col_idx, value=mileage_data[student_id])
        else:
            new_worksheet.cell(row=row_idx, column=current_mileage_col_idx, value="None")
        
        # 设置格式
        new_worksheet.cell(row=row_idx, column=current_mileage_col_idx).style = text_style
        new_worksheet.cell(row=row_idx, column=current_mileage_col_idx).alignment = Alignment(horizontal='center')
    
    # 最后统一处理学号、密码和时间列的格式
    # 获取各列的索引
    headers = [cell.value for cell in new_worksheet[1]]
    student_id_col_idx = headers.index("学号") + 1
    password_col_idx = headers.index("密码") + 1
    time_col_idx = headers.index("时间") + 1
    
    # 统一设置学号列为文本格式
    for row_idx in range(2, max_row + 1):
        cell = new_worksheet.cell(row=row_idx, column=student_id_col_idx)
        if cell.value is not None:
            # 确保学号是文本格式，避免长数字变成科学计数法
            cell.value = str(cell.value)
            cell.style = text_style
            cell.alignment = Alignment(horizontal='center')
    
    # 统一设置密码列为文本格式
    for row_idx in range(2, max_row + 1):
        cell = new_worksheet.cell(row=row_idx, column=password_col_idx)
        if cell.value is not None:
            # 确保密码是文本格式
            cell.value = str(cell.value)
            cell.style = text_style
            cell.alignment = Alignment(horizontal='center')
    
    # 统一设置时间列为日期格式（m-d）
    for row_idx in range(2, max_row + 1):
        cell = new_worksheet.cell(row=row_idx, column=time_col_idx)
        raw_value = cell.value

        if raw_value is None or raw_value == "":
            continue

        # 如果已经是日期/日期时间类型，只需设置显示格式
        if isinstance(raw_value, (datetime, date)):
            cell.number_format = "m-d"
            cell.alignment = Alignment(horizontal="center")
            continue

        text = str(raw_value).strip()

        # 优先处理形如 mm.dd 的数字/文本，例如 9.1、09.01 等
        try:
            parts = text.split(".")
            if len(parts) == 2:
                month = int(parts[0])
                day = int(parts[1])
                year = datetime.now().year
                cell.value = datetime(year, month, day)
                cell.number_format = "m-d"
                cell.alignment = Alignment(horizontal="center")
                continue
        except Exception:
            # 解析失败时退回到纯文本处理
            pass

        # 无法按 mm.dd 解析时，作为文本写入并替换为 mm-dd 形式
        cell.value = text.replace(".", "-")
        cell.style = text_style
        cell.alignment = Alignment(horizontal="center")

    # 保存新文件
    new_workbook.save(output_path)


def ensure_excel_for_computer_updated() -> None:
    """
    确保供程序使用的 Excel 文件已根据原始文件生成/更新。
    - 源文件不存在时静默返回，不中断主流程；
    - 当源文件比目标文件新或目标文件不存在时，重新生成。
    """
    input_file = EXCEL_SOURCE_FILE
    output_file = EXCEL_FILE

    if not input_file.exists():
        return

    if (not output_file.exists()
            or input_file.stat().st_mtime > output_file.stat().st_mtime):
        process_excel_file(str(input_file), str(output_file))


if __name__ == "__main__":
    ensure_excel_for_computer_updated()
    print("处理完成，结果已保存")
