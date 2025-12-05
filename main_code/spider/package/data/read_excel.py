import openpyxl
import json
from typing import List, Dict, Any, Union
import os

# 使用统一的绝对路径配置
from paths import EXCEL_FILE
from .update_excel_for_computer import ensure_excel_for_computer_updated

# ---------- 获取字段到列索引的映射 ----------
def get_field_to_column_mapping(file_path: str) -> Dict[str, int]:
    """
    获取字段名到列索引的映射
    返回一个字典，键是字段名，值是对应的列索引（从1开始）
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Sheet1"]

    # 获取表头
    headers = [cell.value for cell in ws[1]]

    # 创建字段名到列索引的映射
    field_to_column = {}
    for idx, header in enumerate(headers, 1):
        if header is not None:
            field_to_column[str(header)] = idx

    return field_to_column

# ---------- 主要数据提取函数 ----------
def extract_data(fields=None, default_rules=None):
    """
    从2025.9.1_for_computer.xlsx提取数据
    使用字段名而不是列索引，增强健壮性

    参数:
        fields: 要提取的字段名列表，例如 ['学号', '密码', '当前里程']
               如果为None，则提取所有字段
        default_rules: 默认规则字典，格式为 {字段名: ('func'|'value', 值)}

    返回:
        以学号为键，字段值列表为值的字典
    """
    # 每次读取前，先确保供程序使用的 Excel 文件已经根据原始文件更新（包括时间字段格式）
    ensure_excel_for_computer_updated()

    # 使用绝对路径
    file_path = str(EXCEL_FILE)
    
    # 获取字段名到列索引的映射
    field_to_column = get_field_to_column_mapping(file_path)
    
    # 如果没有指定字段，则提取所有字段
    if fields is None:
        fields = list(field_to_column.keys())
    elif isinstance(fields, str):
        fields = [fields]
    
    # 设置默认规则
    if default_rules is None:
        default_rules = {}
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Sheet1"]
    result = {}

    for row in range(2, ws.max_row + 1):
        account = ws.cell(row=row, column=field_to_column.get("学号", 1)).value
        if account is None:
            break
        account_key = str(account)

        attrs = []
        for field in fields:
            col_idx = field_to_column.get(field)
            if col_idx is None:
                # 如果字段不存在，尝试使用默认规则
                if field in default_rules:
                    rule_type, rule_val = default_rules[field]
                    value = rule_val(account_key, None) if rule_type == 'func' else rule_val
                else:
                    value = None
            else:
                value = ws.cell(row=row, column=col_idx).value
                # 如果值为空且字段有默认规则，使用默认规则
                if value is None and field in default_rules:
                    rule_type, rule_val = default_rules[field]
                    value = rule_val(account_key, value) if rule_type == 'func' else rule_val
            
            attrs.append(value)
        result[account_key] = attrs

    return result

if __name__ == '__main__':
    # 示例：使用字段名而不是列索引
    print(extract_data(["学号", "密码", "当前里程"]))
    
    # 示例：获取需要线上学习或考试的学生数据
    all_data = extract_data(["学号", "密码", "线上学习", "考试"])
    online_exam_data = {k: [v[1], int(v[2] or 0), int(v[3] or 0)] 
                        for k, v in all_data.items() 
                        if v[2] == 1 or v[3] == 1}
    print(online_exam_data)
